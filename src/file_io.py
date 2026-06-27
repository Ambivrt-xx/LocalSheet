"""
File I/O — open/save XLSX (with formatting) and CSV using openpyxl + csv.
"""
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

from .models import Cell, CellFormat, Worksheet, Workbook, ConditionalRule


# ───────────────────── helpers ─────────────────────

def _argb_to_hex(argb):
    """Convert openpyxl ARGB color string to '#RRGGBB'."""
    if argb is None:
        return None
    s = str(argb)
    if s == '00000000':
        return None
    if len(s) == 8:
        return '#' + s[2:]
    if len(s) == 6:
        return '#' + s
    return None


def _hex_to_argb(hex_color):
    """Convert '#RRGGBB' to openpyxl ARGB 'FFRRGGBB'."""
    if hex_color is None:
        return None
    h = hex_color.lstrip('#')
    if len(h) == 6:
        return 'FF' + h
    return None


# ───────────────────── Save XLSX ─────────────────────

def save_xlsx(workbook, file_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet in workbook.sheets:
        ws = wb.create_sheet(title=sheet.name[:31])

        # tab color
        if sheet.tab_color:
            ws.sheet_properties.tabColor = sheet.tab_color.lstrip('#')

        # sheet visibility
        if sheet.sheet_hidden:
            ws.sheet_state = 'hidden'

        # gridlines
        if not sheet.show_gridlines:
            ws.sheet_view.showGridLines = False

        # column widths
        for col, width in sheet.col_widths.items():
            ws.column_dimensions[get_column_letter(col + 1)].width = max(width / 7, 8)

        # row heights
        for row, height in sheet.row_heights.items():
            ws.row_dimensions[row + 1].height = height * 0.75

        # cells
        for (row, col), cell in sheet.cells.items():
            if cell.raw == "" and not cell.fmt.bold and not cell.fmt.italic and \
               not cell.fmt.underline and cell.fmt.font_family is None and \
               cell.fmt.text_color is None and cell.fmt.bg_color is None and \
               cell.fmt.align is None and cell.fmt.border is None:
                continue

            ec = ws.cell(row=row + 1, column=col + 1)

            if cell.is_formula():
                ec.value = cell.raw
            elif cell.raw != "":
                ec.value = cell.value
            else:
                continue

            fmt = cell.fmt

            # font
            font_kwargs = {}
            if fmt.bold:
                font_kwargs['bold'] = True
            if fmt.italic:
                font_kwargs['italic'] = True
            if fmt.underline:
                font_kwargs['underline'] = True
            if fmt.strikethrough:
                font_kwargs['strike'] = True
            if fmt.font_family:
                font_kwargs['name'] = fmt.font_family
            if fmt.font_size and fmt.font_size != 11:
                font_kwargs['size'] = fmt.font_size
            if fmt.text_color:
                font_kwargs['color'] = _hex_to_argb(fmt.text_color)
            if font_kwargs:
                ec.font = Font(**font_kwargs)

            # fill
            if fmt.bg_color:
                ec.fill = PatternFill(fill_type='solid',
                                     fgColor=_hex_to_argb(fmt.bg_color))

            # alignment
            align_kwargs = {}
            if fmt.align:
                align_kwargs['horizontal'] = fmt.align
            if fmt.wrap_text:
                align_kwargs['wrap_text'] = True
            if fmt.indent:
                align_kwargs['indent'] = fmt.indent
            if fmt.text_rotation:
                align_kwargs['text_rotation'] = fmt.text_rotation if fmt.text_rotation != 255 else 255
            if align_kwargs:
                ec.alignment = Alignment(**align_kwargs)

            # border
            if fmt.border:
                side = Side(style='thin')
                if fmt.border == 'all' or fmt.border == 'grid':
                    ec.border = Border(left=side, right=side, top=side, bottom=side)
                elif fmt.border == 'outline':
                    ec.border = Border(left=side, right=side, top=side, bottom=side)
                elif fmt.border == 'bottom':
                    ec.border = Border(bottom=side)
                elif fmt.border == 'top':
                    ec.border = Border(top=side)
                elif fmt.border == 'left':
                    ec.border = Border(left=side)
                elif fmt.border == 'right':
                    ec.border = Border(right=side)

            # cell note (comment)
            if cell.note:
                from openpyxl.comments import Comment
                ec.comment = Comment(cell.note, "LocalSheet")

            # number format
            if fmt.custom_format:
                ec.number_format = fmt.custom_format
            elif fmt.number_format:
                nf_map = {
                    'currency': '"$"#,##0.00',
                    'percent': '0%',
                    'float2': '0.00',
                    'int': '0',
                    'thousands': '#,##0',
                    'date': 'mm/dd/yyyy',
                    'time': 'hh:mm:ss',
                    'datetime': 'mm/dd/yyyy hh:mm',
                }
                ec.number_format = nf_map.get(fmt.number_format, 'General')

            # hyperlink
            if cell.hyperlink:
                ec.hyperlink = cell.hyperlink

        # conditional formatting
        for rule in getattr(sheet, 'conditional_rules', []):
            range_str = f"{get_column_letter(rule.left + 1)}{rule.top + 1}:{get_column_letter(rule.right + 1)}{rule.bottom + 1}"
            if rule.rule_type == 'color_scale':
                cs = ColorScaleRule(
                    start_type='min', start_color=rule.min_color.lstrip('#'),
                    mid_type='percentile', mid_value=50, mid_color=rule.mid_color.lstrip('#'),
                    end_type='max', end_color=rule.max_color.lstrip('#'))
                ws.conditional_formatting.add(range_str, cs)
            elif rule.rule_type == 'data_bar':
                db = DataBarRule(start_type='min', end_type='max',
                                 color=rule.bar_color.lstrip('#'))
                ws.conditional_formatting.add(range_str, db)

        # freeze panes
        if sheet.frozen_rows > 0 or sheet.frozen_cols > 0:
            freeze_col = get_column_letter(sheet.frozen_cols + 1)
            freeze_row = sheet.frozen_rows + 1
            ws.freeze_panes = f"{freeze_col}{freeze_row}"

        # merged cells
        for (r1, c1, r2, c2) in sheet.merged_cells:
            ws.merge_cells(start_row=r1+1, start_column=c1+1,
                          end_row=r2+1, end_column=c2+1)

        # data validations
        from openpyxl.worksheet.datavalidation import DataValidation as OpDataValidation
        for (r, c), dv in sheet.data_validations.items():
            if dv.dv_type == 'list':
                odv = OpDataValidation(type='list', formula1=f'"{dv.formula1}"',
                                       allow_blank=dv.allow_blank)
                if dv.prompt:
                    odv.prompt = dv.prompt
                    odv.promptTitle = "Input"
                odv.add(f"{get_column_letter(c+1)}{r+1}")
                ws.add_data_validation(odv)

        # named ranges
        for name, (r1, c1, r2, c2) in sheet.named_ranges.items():
            ref = f"{ws.title}!${get_column_letter(c1+1)}${r1+1}:${get_column_letter(c2+1)}${r2+1}"
            wb.defined_names[name] = openpyxl.workbook.defined_name.DefinedName(name=name, attr_text=ref)

    wb.save(file_path)


# ───────────────────── Load XLSX ─────────────────────

def load_xlsx(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    workbook = Workbook()
    workbook.sheets = []
    workbook.active_index = 0

    for ws in wb.worksheets:
        sheet = Worksheet(ws.title[:31])

        # tab color
        if ws.sheet_properties.tabColor:
            sheet.tab_color = _argb_to_hex(str(ws.sheet_properties.tabColor))

        # sheet visibility
        if ws.sheet_state == 'hidden':
            sheet.sheet_hidden = True

        # gridlines
        if ws.sheet_view and not ws.sheet_view.showGridLines:
            sheet.show_gridlines = False

        # column widths
        for col_letter, dim in ws.column_dimensions.items():
            if dim.width:
                # convert openpyxl width to pixels (approx)
                col_idx = openpyxl.utils.column_index_from_string(col_letter) - 1
                sheet.col_widths[col_idx] = int(dim.width * 7)

        # row heights
        for row_num, dim in ws.row_dimensions.items():
            if dim.height:
                sheet.row_heights[row_num - 1] = int(dim.height / 0.75)

        # cells
        for row in ws.iter_rows():
            for ec in row:
                if ec.value is None:
                    # check formatting only
                    if ec.font.bold or ec.font.italic or ec.fill.fgColor or ec.alignment.horizontal:
                        cell = sheet.get_cell(ec.row - 1, ec.column - 1)
                        cell.raw = ""
                        cell.value = ""
                    continue

                r, c = ec.row - 1, ec.column - 1
                cell = sheet.get_cell(r, c)

                if isinstance(ec.value, str) and ec.value.startswith('='):
                    cell.raw = ec.value
                    cell.value = None  # will be recalculated
                else:
                    cell.raw = str(ec.value) if not isinstance(ec.value, (int, float)) else str(ec.value)
                    cell.value = ec.value

                # formatting
                fmt = cell.fmt
                if ec.font:
                    fmt.bold = bool(ec.font.bold)
                    fmt.italic = bool(ec.font.italic)
                    fmt.underline = bool(ec.font.underline)
                    if ec.font.strike:
                        fmt.strikethrough = True
                    if ec.font.name:
                        fmt.font_family = ec.font.name
                    if ec.font.size:
                        fmt.font_size = int(ec.font.size)
                    color = ec.font.color
                    if color and color.rgb:
                        fmt.text_color = _argb_to_hex(str(color.rgb))
                if ec.fill and ec.fill.fill_type == 'solid':
                    fg = ec.fill.fgColor
                    if fg and fg.rgb:
                        fmt.bg_color = _argb_to_hex(str(fg.rgb))
                if ec.alignment:
                    if ec.alignment.horizontal:
                        fmt.align = ec.alignment.horizontal
                    if ec.alignment.wrap_text:
                        fmt.wrap_text = True
                    if ec.alignment.indent:
                        fmt.indent = int(ec.alignment.indent)
                    if ec.alignment.text_rotation:
                        fmt.text_rotation = int(ec.alignment.text_rotation)
                # number format
                nf = ec.number_format
                if nf:
                    if nf in ('0%', '0.0%', '0.00%'):
                        fmt.number_format = 'percent'
                    elif nf in ('0', '#,##0'):
                        fmt.number_format = 'thousands' if ',' in nf else 'int'
                    elif nf in ('0.00', '#,##0.00'):
                        fmt.number_format = 'float2'
                    elif '$' in nf or 'Currency' in nf:
                        fmt.number_format = 'currency'
                    elif nf in ('mm/dd/yyyy', 'm/d/yyyy', 'yyyy-mm-dd', 'dd/mm/yyyy'):
                        fmt.number_format = 'date'
                    elif nf in ('hh:mm:ss', 'hh:mm', 'h:mm:ss'):
                        fmt.number_format = 'time'
                    elif 'mm/dd' in nf and 'hh' in nf:
                        fmt.number_format = 'datetime'
                    elif nf not in ('General', 'general'):
                        fmt.custom_format = nf
                # hyperlink
                if ec.hyperlink:
                    cell.hyperlink = ec.hyperlink.target
                if ec.border:
                    has_all = all([
                        ec.border.left and ec.border.left.style,
                        ec.border.right and ec.border.right.style,
                        ec.border.top and ec.border.top.style,
                        ec.border.bottom and ec.border.bottom.style,
                    ])
                    if has_all:
                        fmt.border = 'all'
                    elif ec.border.bottom and ec.border.bottom.style:
                        fmt.border = 'bottom'
                    elif ec.border.top and ec.border.top.style:
                        fmt.border = 'top'

        # conditional formatting
        for cf in ws.conditional_formatting:
            try:
                sqref_str = str(cf.sqref)
                min_col, min_row, max_col, max_row = range_boundaries(sqref_str)
            except Exception:
                continue
            top, left = min_row - 1, min_col - 1
            bottom, right = max_row - 1, max_col - 1
            for rule in cf.rules:
                if rule.type == 'colorScale' and rule.colorScale:
                    colors = rule.colorScale.color
                    cfvos = rule.colorScale.cfvo
                    min_c = _argb_to_hex(str(colors[0].rgb)) if colors and len(colors) > 0 else '#f8696b'
                    mid_c = _argb_to_hex(str(colors[1].rgb)) if colors and len(colors) > 1 else '#ffeb84'
                    max_c = _argb_to_hex(str(colors[2].rgb)) if colors and len(colors) > 2 else '#63be7b'
                    sheet.conditional_rules.append(ConditionalRule(
                        rule_type='color_scale', top=top, left=left,
                        bottom=bottom, right=right,
                        min_color=min_c, mid_color=mid_c, max_color=max_c))
                elif rule.type == 'dataBar' and rule.dataBar:
                    colors = rule.dataBar.color
                    bar_c = '#638ec6'
                    if colors and len(colors) > 0:
                        bar_c = _argb_to_hex(str(colors[0].rgb)) or '#638ec6'
                    sheet.conditional_rules.append(ConditionalRule(
                        rule_type='data_bar', top=top, left=left,
                        bottom=bottom, right=right, bar_color=bar_c))

        # freeze panes
        if ws.freeze_panes:
            fp = ws.freeze_panes
            if isinstance(fp, str):
                col_part = ''.join(ch for ch in fp if ch.isalpha())
                row_part = ''.join(ch for ch in fp if ch.isdigit())
                if col_part:
                    sheet.frozen_cols = openpyxl.utils.column_index_from_string(col_part) - 1
                if row_part:
                    sheet.frozen_rows = int(row_part) - 1

        # merged cells
        for merged_range in ws.merged_cells.ranges:
            try:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
                sheet.merged_cells.append((min_row-1, min_col-1, max_row-1, max_col-1))
            except Exception:
                pass

        # data validations
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list' and dv.formula1:
                values = dv.formula1.strip('"')
                for cell_range in dv.sqref.ranges:
                    for row in range(cell_range.min_row, cell_range.max_row + 1):
                        for col in range(cell_range.min_col, cell_range.max_col + 1):
                            from .models import DataValidation
                            sheet.data_validations[(row-1, col-1)] = DataValidation(
                                dv_type='list', formula1=values,
                                allow_blank=dv.allow_blank,
                                prompt=dv.prompt or '')

        # named ranges
        for name in wb.defined_names:
            try:
                dn = wb.defined_names[name]
                ref = dn.attr_text
                # Parse "Sheet1!$A$1:$B$2"
                import re
                m = re.match(r'(?:[^!]+!)?\$?([A-Za-z]+)\$?(\d+)(?::\$?([A-Za-z]+)\$?(\d+))?', ref)
                if m:
                    c1 = openpyxl.utils.column_index_from_string(m.group(1)) - 1
                    r1 = int(m.group(2)) - 1
                    if m.group(3):
                        c2 = openpyxl.utils.column_index_from_string(m.group(3)) - 1
                        r2 = int(m.group(4)) - 1
                    else:
                        c2, r2 = c1, r1
                    sheet.named_ranges[name] = (r1, c1, r2, c2)
            except Exception:
                pass

        workbook.sheets.append(sheet)

    if not workbook.sheets:
        workbook.add_sheet("Sheet1")

    return workbook


# ───────────────────── CSV ─────────────────────

def export_csv(sheet, file_path):
    if sheet.cells:
        max_row = max(r for r, _ in sheet.cells) + 1
        max_col = max(c for _, c in sheet.cells) + 1
    else:
        max_row = max_col = 0

    with open(file_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        for r in range(max_row):
            row_data = []
            for c in range(max_col):
                cell = sheet.cells.get((r, c))
                row_data.append(cell.display_value() if cell else "")
            writer.writerow(row_data)


def load_csv(file_path):
    workbook = Workbook()
    workbook.sheets = []
    sheet = Worksheet("Sheet1")

    with open(file_path, 'r', newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        for r, row in enumerate(reader):
            for c, val in enumerate(row):
                if val != "":
                    cell = sheet.get_cell(r, c)
                    cell.raw = val
                    from .formula_engine import parse_value
                    cell.value = parse_value(val)

    workbook.sheets.append(sheet)
    workbook.active_index = 0
    return workbook